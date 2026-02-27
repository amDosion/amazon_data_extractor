"""
Excel文件处理工具类
提供统一的Excel生成、导出功能
"""
import io
import tempfile
from typing import List, Dict, Any, Optional
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd


class ExcelExporter:
    """Excel导出工具类"""
    
    @staticmethod
    def create_template(
        headers: List[str],
        sample_data: List[List[Any]] = None,
        sheet_name: str = "Template",
        filename: str = "template.xlsx",
        dropdown_columns: Dict[str, List[str]] = None
    ) -> FileResponse:
        """
        创建Excel模板文件
        
        Args:
            headers: 表头列表
            sample_data: 示例数据（可选）
            sheet_name: 工作表名称
            filename: 文件名
            
        Returns:
            FileResponse: Excel文件响应
        """
        try:
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 设置表头
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 添加示例数据
            if sample_data:
                for row_data in sample_data:
                    ws.append(row_data)
            
            # 设置列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 添加下拉列表数据验证
            if dropdown_columns:
                for header, options in dropdown_columns.items():
                    try:
                        # 找到对应列的索引
                        col_index = headers.index(header) + 1
                        col_letter = ws.cell(row=1, column=col_index).column_letter
                        
                        # 创建下拉列表验证
                        options_str = ','.join(options)
                        dv = DataValidation(
                            type="list",
                            formula1=f'"{options_str}"',
                            allow_blank=True
                        )
                        dv.error = f'请选择有效的{header}'
                        dv.errorTitle = '输入错误'
                        dv.prompt = f'请从列表中选择{header}'
                        dv.promptTitle = '选择提示'
                        
                        # 应用到整列（从第2行开始，前1000行）
                        ws.add_data_validation(dv)
                        dv.add(f'{col_letter}2:{col_letter}1000')
                        
                    except ValueError:
                        # 如果header不在列表中，跳过
                        continue
            
            # 保存文件
            wb.save(temp_file.name)
            
            # 返回文件响应
            return FileResponse(
                path=temp_file.name,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
        except Exception as e:
            raise Exception(f"创建Excel模板失败: {str(e)}")
    
    @staticmethod
    def export_data(
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        filename: str = None
    ) -> FileResponse:
        """
        导出数据到Excel文件
        
        Args:
            data: 要导出的数据列表
            sheet_name: 工作表名称
            filename: 文件名（如果为None，自动生成）
            
        Returns:
            FileResponse: Excel文件响应
        """
        try:
            if not data:
                raise Exception("没有数据可导出")
            
            # 生成文件名
            if not filename:
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"export_data_{timestamp}.xlsx"
            
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            
            # 使用pandas创建Excel
            df = pd.DataFrame(data)
            with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表并设置样式
                worksheet = writer.sheets[sheet_name]
                
                # 设置表头样式
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 返回文件响应
            return FileResponse(
                path=temp_file.name,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"
                }
            )
        except Exception as e:
            raise Exception(f"导出Excel文件失败: {str(e)}")


class ExcelParser:
    """Excel解析工具类"""
    
    @staticmethod
    def parse_upload_file(file_content: bytes, required_columns: List[str] = None) -> pd.DataFrame:
        """
        解析上传的Excel文件
        
        Args:
            file_content: 文件内容字节
            required_columns: 必需的列名列表
            
        Returns:
            pd.DataFrame: 解析后的数据框
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content))
            
            # 验证必需的列
            if required_columns:
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    raise Exception(f"缺少必需的列: {', '.join(missing_columns)}")
            
            return df
        except Exception as e:
            raise Exception(f"解析Excel文件失败: {str(e)}")
    
    @staticmethod
    def clean_cell_value(value: Any) -> Optional[str]:
        """
        清理单元格值，处理空值和NaN
        
        Args:
            value: 单元格原始值
            
        Returns:
            Optional[str]: 清理后的字符串值或None
        """
        if pd.isna(value) or str(value).strip() == '' or str(value).lower() == 'nan':
            return None
        return str(value).strip()
